from fastapi import FastAPI, Query, UploadFile, File, Request, HTTPException
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import gspread
from google.oauth2.service_account import Credentials
import re, os, io, threading, uuid, gc, time
import concurrent.futures
from datetime import timedelta, datetime
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote

# LINE Bot 官方套件
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage, FileMessage

app = FastAPI(title="🚗 杰運汽車新竹店阿鍇專用 - 內部系統")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

LINE_CHANNEL_ACCESS_TOKEN = "Vetc+mW1cmCmkEkXI7GcWpVtqqCkSEDSp/wQuOrQB0SA2GCanyXBmMczQzRW+CK8Obpv2gOMap4rtxRQIa/8/8eqCpdBm/zwozhJndUIEe+NSwPITjCVkPDbKG3usLC/jkh8KlqEkbDoAM8XFYTLRwdB04t89/1O/w1cDnyilFU="
LINE_CHANNEL_SECRET = "ff5426c6ab3102189f8d45f0eca69652"

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

SHEET_ID = "1HWb5u6EGYSHVJHFhmhmsVv4xDgHlQEkdicfXBuFp86w"
CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=0"
SIMPLE_CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=852175657"

cached_df = None
view_api_session = None
cached_valid_u = None
cached_valid_p = None

KNOWN_MAKES = [
    "TOYOTA", "HONDA", "BENZ", "BMW", "AUDI", "LEXUS", "VOLVO", "VW", "MAZDA", 
    "NISSAN", "FORD", "PORSCHE", "MG", "SKODA", "MINI", "KIA", "SUZUKI", 
    "MITSUBISHI", "LUXGEN", "LAND ROVER", "JAGUAR", "SUBARU", "TESLA", 
    "MASERATI", "FERRARI", "LAMBORGHINI", "BENTLEY", "ROLLS-ROYCE"
]

# 💡 核心引擎：60秒極速快取
class TTLCache:
    def __init__(self, ttl=60):
        self.ttl = ttl
        self.cache = {}
        self.timestamps = {}
        self.lock = threading.Lock()
        
    def get(self, key):
        with self.lock:
            if key in self.cache and time.time() - self.timestamps.get(key, 0) < self.ttl:
                return self.cache[key]
            return None
            
    def set(self, key, value):
        with self.lock:
            self.cache[key] = value
            self.timestamps[key] = time.time()
            
    def clear(self, key):
        with self.lock:
            if key in self.cache:
                del self.cache[key]
                del self.timestamps[key]

api_cache = TTLCache(ttl=60)

def get_gspread_client():
    key_path = "/etc/secrets/google_key.json"
    if not os.path.exists(key_path):
        raise Exception("尚未設定 Google API 憑證！")
    creds = Credentials.from_service_account_file(
        key_path, 
        scopes=['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    )
    return gspread.authorize(creds)

def get_or_create_creds():
    cached = api_cache.get("sys_creds")
    if cached: 
        return cached
    try:
        ws = get_gspread_client().open_by_key(SHEET_ID).worksheet("系統設定")
        data = ws.get_all_values()
        user = data[1][1] if len(data) > 1 and len(data[1]) > 1 else "Admin02"
        pwd = data[2][1] if len(data) > 2 and len(data[2]) > 1 else "Eric740625"
        api_cache.set("sys_creds", (user, pwd))
        return user, pwd
    except Exception:
        return "Admin02", "Eric740625"

def update_creds(user, pwd):
    try:
        get_gspread_client().open_by_key(SHEET_ID).worksheet("系統設定").update(values=[[user], [pwd]], range_name='B2:B3')
    except Exception:
        pass

def get_backup_credentials_from_sheet():
    cached = api_cache.get("bkp_creds")
    if cached is not None: 
        return cached
    try:
        raw_data = get_gspread_client().open_by_key(SHEET_ID).worksheet("員工編號列表").get_all_values()
        b_creds = []
        for row in raw_data[1:]:
            if len(row) > 1 and str(row[1]).strip():
                b_creds.append((str(row[1]).strip(), str(row[2]).strip() if len(row) > 2 else "123456"))
        api_cache.set("bkp_creds", b_creds)
        return b_creds
    except Exception:
        return []

def check_permission(user_id, action):
    if not user_id: 
        return False
    records = api_cache.get("perm_records")
    if records is None:
        try:
            raw_data = get_gspread_client().open_by_key(SHEET_ID).worksheet("權限管理").get_all_values()
            if not raw_data: 
                return False
            records = [dict(zip(raw_data[0], row)) for row in raw_data[1:]]
            api_cache.set("perm_records", records)
        except Exception:
            return False
            
    tw_now = datetime.utcnow() + timedelta(hours=8)
    for r in records:
        if str(r.get("LINE ID", "")).strip() == str(user_id).strip():
            is_super = str(r.get("最高管理員", "")).strip().upper() == "V"
            exp_str = str(r.get("到期日", "")).strip()
            
            if exp_str and not is_super:
                try:
                    exp_date = datetime.strptime(exp_str.replace("-", "/"), "%Y/%m/%d") + timedelta(days=1)
                    if tw_now >= exp_date: 
                        return False 
                except Exception:
                    pass
                    
            if is_super: 
                return True
            return str(r.get(action, "")).strip().upper() == "V"
            
    return False

def clean_money(val):
    if pd.isna(val): return 0.0
    s = str(val).replace(',', '')
    matches = re.findall(r"(\d+\.?\d*)", s)
    if matches:
        try:
            v = float(matches[-1])
            return round(v / 10000, 2) if v > 1000 else v
        except Exception:
            return 0.0
    return 0.0

def parse_roc_date(date_val):
    if pd.isna(date_val): return pd.NaT
    s = str(date_val).strip().replace(".", "/").replace("-", "/")
    if not s: return pd.NaT
    try:
        parts = s.split('/')
        if len(parts) == 3:
            y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            return pd.Timestamp(year=y + 1911 if y < 1911 else y, month=m, day=d)
        return pd.to_datetime(s, errors='coerce')
    except Exception:
        return pd.NaT

def load_and_clean_data():
    global cached_df
    client = get_gspread_client()
    doc = client.open_by_key(SHEET_ID)
    dfs = []
    
    try:
        ws_main = doc.worksheet("E車源")
        df_main = pd.read_csv(f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={ws_main.id}")
        if "車牌" not in df_main.columns:
            for idx, row in df_main.iterrows():
                vals = [str(x).strip() for x in row.values]
                if "車牌" in vals and ("廠牌" in vals or "品牌" in vals or "車型" in vals):
                    df_main.columns = vals
                    df_main = df_main.iloc[idx+1:].reset_index(drop=True)
                    break
        df_main['is_sold_sheet'] = False
        dfs.append(df_main)
    except Exception:
        pass

    if not dfs: 
        df = pd.read_csv(CSV_URL)
        if "車牌" not in df.columns:
            for idx, row in df.iterrows():
                vals = [str(x).strip() for x in row.values]
                if "車牌" in vals:
                    df.columns = vals
                    df = df.iloc[idx+1:].reset_index(drop=True)
                    break
        df['is_sold_sheet'] = False
    else: 
        df = pd.concat(dfs, ignore_index=True)

    df.columns = [str(c).strip() for c in df.columns]
    
    if '採購' not in df.columns: 
        df['採購'] = df.get('採購人', df.get('車輛負責人', df.get('負責人', "")))

    df['編號'] = df.apply(
        lambda r: f"{str(r.get('舊編號','')).replace('.0','')} ({str(r.get('新編號','')).replace('.0','')})" 
        if str(r.get('新編號','')).strip() and str(r.get('舊編號','')).strip() 
        else (str(r.get('新編號','')) or str(r.get('舊編號',''))), 
        axis=1
    )
    
    price_col = '網路' if '網路' in df.columns else ('售價' if '售價' in df.columns else ('價格' if '價格' in df.columns else '底價'))
    df['顯示價格'] = df[price_col].apply(clean_money) if price_col in df.columns else 0.0

    brand_col = '廠牌' if '廠牌' in df.columns else ('品牌' if '品牌' in df.columns else None)
    if brand_col: 
        df['廠牌'] = df[brand_col].apply(lambda b: re.sub(r'[\u4e00-\u9fa5]', '', str(b).split('/')[0]).strip().upper())
        
    if '年份' in df.columns: 
        df['年份'] = df['年份'].astype(str)
        
    if '里程' in df.columns: 
        df['里程'] = df['里程'].apply(
            lambda m: "" if pd.isna(m) or str(m).strip().lower() == 'nan' 
            else (str(m).replace(',', '').strip()[:-2] if str(m).replace(',', '').strip().endswith('.0') else str(m).replace(',', '').strip())
        )

    if '車輛位置' in df.columns:
        def clean_loc(loc):
            loc = str(loc).strip()
            if '台北' in loc or '北投' in loc: return '北投店'
            if '桃園' in loc: return '桃園店'
            if '台中' in loc: return '台中店'
            if '高雄' in loc: return '高雄新廠'
            return loc
        df['車輛位置'] = df['車輛位置'].apply(clean_loc)

    def normalize_property(row):
        full = str(row.get('產權', '')) + str(row.get('公司', ''))
        if '禾迪' in full: return '禾迪'
        if '展帆' in full: return '展帆'
        if '租車' in full or '杰租' in full: return '杰租'
        return '杰運' if '杰' in full else '其他'
        
    df['filter_property'] = df.apply(normalize_property, axis=1)
    
    if '狀態' in df.columns:
        df['is_sold'] = df.apply(lambda r: True if '已售' in str(r.get('狀態', '')) or r.get('is_sold_sheet') else False, axis=1)
        df['is_cert'] = df['狀態'].apply(lambda x: True if '取證' in str(x) else False)
    else:
        df['is_sold'] = df.get('is_sold_sheet', False)
        df['is_cert'] = False
        
    df['is_reserved'] = df.apply(lambda r: True if '已收訂' in str(r.get('狀態', '')) or '已收訂' in str(r.get('收訂狀態', '')) else False, axis=1)
    
    date_col = '入庫日期' if '入庫日期' in df.columns else ('入庫日' if '入庫日' in df.columns else None)
    if date_col: 
        df['入庫_dt'] = df[date_col].apply(parse_roc_date)
    
    cached_df = df.fillna("")
    gc.collect() 
    return cached_df

def process_crm_excel(filename: str, contents: bytes):
    wb = None
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        
        new_customers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            r_dict = {headers[i]: str(row[i]).strip() if row[i] is not None else "" for i in range(min(len(headers), len(row)))}
            name = r_dict.get("姓名", "")
            if not name: continue
                
            phone = ""
            for k, v in r_dict.items():
                if ("手機" in k or "電話" in k) and v:
                    clean_p = re.sub(r'\D', '', v)
                    if clean_p.startswith("09") and len(clean_p) >= 10: 
                        phone = clean_p[:10]
                        break
            if not phone: continue 
                
            date_val = r_dict.get("生效日", "") or r_dict.get("建立時間", "")
            memo = r_dict.get("附註", "")
            tags = r_dict.get("標籤", "")
            status = "新客詢問"
            if "成交" in tags: status = "已成交"
            elif "收訂" in tags: status = "已收訂"
            elif "戰敗" in tags or "放棄" in tags or "暫緩" in tags: status = "戰敗"
            elif "賞車" in tags or "看車" in tags: status = "安排賞車"
            
            sales = next((v for k, v in r_dict.items() if ("業務" in k or "負責" in k) and v and "@" not in v), r_dict.get("客戶擴充欄位-銷售業務", ""))
            needs = next((v for k, v in r_dict.items() if ("車" in k or "需求" in k) and "車牌" not in k and "車身" not in k and v), "")
            
            new_customers.append({
                "日期": date_val, "姓名": name, "電話": f"'{phone}", 
                "需求車款": needs, "負責業務": sales, "狀態": status, "備註": memo
            })
            
        client = get_gspread_client()
        doc = client.open_by_key(SHEET_ID)
        try:
            sheet = doc.worksheet("客資紀錄")
            raw_values = sheet.get_all_values()
            old_records = []
            if raw_values and len(raw_values) > 1:
                hdrs = [str(h).strip() for h in raw_values[0]]
                for r in raw_values[1:]:
                    padded = list(r)
                    while len(padded) < len(hdrs): padded.append("")
                    old_records.append(dict(zip(hdrs, padded)))
        except Exception:
            sheet = doc.add_worksheet("客資紀錄", 1000, 10)
            old_records = []
            
        merged_dict = {str(rec.get("電話", "")).replace("'", "").strip(): rec for rec in old_records if str(rec.get("電話", "")).replace("'", "").strip()}
        update_count, add_count = 0, 0
        
        for nc in new_customers:
            p = nc["電話"].replace("'", "")
            if p in merged_dict:
                update_count += 1
                merged_dict[p].update({k: v for k, v in nc.items() if v and (k != "狀態" or v != "新客詢問")})
            else:
                add_count += 1
                merged_dict[p] = nc
                
        headers_crm = ["日期", "姓名", "電話", "需求車款", "負責業務", "狀態", "備註"]
        final_data = [headers_crm] + [[str(rec.get(h, "")) for h in headers_crm] for rec in merged_dict.values()]
        sheet.clear()
        sheet.update(values=final_data, range_name="A1")
        return {"status": "success", "message": f"👥 客資同步完成！\n本次新增 {add_count} 筆，更新 {update_count} 筆。"}
    except Exception as e: 
        return {"status": "error", "message": f"客資處理失敗：{str(e)}"}
    finally:
        if wb: wb.close()
        gc.collect()

def process_pdf_file(filename: str, contents: bytes):
    try: 
        import pdfplumber
    except ImportError: 
        return {"status": "error", "message": "缺少 pdfplumber 套件"}
        
    try:
        target_tab_name = "新竹車源" if "新竹" in filename else "E車源"
        client = get_gspread_client()
        doc = client.open_by_key(SHEET_ID)
        
        try: 
            target_gsheet = doc.worksheet(target_tab_name)
        except Exception: 
            return {"status": "error", "message": f"找不到 {target_tab_name}"}
            
        all_rows, headers = [], []
        with pdfplumber.open(io.BytesIO(contents)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    for row in table:
                        cleaned_row = [str(cell).replace('\n', ' ').strip() if cell is not None else "" for cell in row]
                        if not any(cleaned_row): continue
                        if not headers and any(kw in str(cleaned_row) for kw in ["車牌", "廠牌", "品牌", "年份", "新編號"]):
                            headers = cleaned_row
                            continue
                        if headers: 
                            all_rows.append(cleaned_row)
                            
        if not headers: 
            return {"status": "error", "message": "無法解析 PDF 表格"}
            
        if "狀態" not in headers: headers.append("狀態")
        status_col_idx = headers.index("狀態")
        data_to_upload = [headers]
        
        for row in all_rows:
            while len(row) <= status_col_idx: row.append("")
            row[status_col_idx] = "在庫"
            data_to_upload.append(row)
            
        target_gsheet.clear()
        target_gsheet.update(values=[[str(cell) for cell in row] for row in data_to_upload], range_name='A1')
        load_and_clean_data()
        return {"status": "success", "message": f"📄 PDF 解析成功！共更新 {len(data_to_upload)-1} 筆車輛！"}
    except Exception as e: 
        return {"status": "error", "message": f"PDF 處理失敗：{str(e)}"}
    finally: 
        gc.collect()

def get_color_rgb(cell):
    try:
        fill = cell.fill
        if not fill: return None
        color = getattr(fill, 'fgColor', None) or getattr(fill, 'start_color', None)
        if not color: return None
        
        rgb_hex = None
        if hasattr(color, 'rgb') and color.rgb and isinstance(color.rgb, str): 
            rgb_hex = color.rgb
        elif hasattr(color, 'type') and color.type == 'indexed':
            from openpyxl.styles.colors import COLOR_INDEX
            idx = color.indexed
            if isinstance(idx, int) and idx < len(COLOR_INDEX): rgb_hex = COLOR_INDEX[idx]
        elif hasattr(color, 'type') and color.type == 'theme':
            theme_colors = ["FFFFFF", "000000", "E7E6E6", "44546A", "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47"]
            idx = color.theme
            if isinstance(idx, int) and idx < len(theme_colors): rgb_hex = theme_colors[idx]
            
        if rgb_hex and isinstance(rgb_hex, str):
            rgb_hex = rgb_hex.replace('#', '')
            if rgb_hex in ['00000000', 'FFFFFFFF']: return None
            if len(rgb_hex) == 8: rgb_hex = rgb_hex[2:] 
            if len(rgb_hex) == 6: 
                return (int(rgb_hex[0:2], 16) / 255.0, int(rgb_hex[2:4], 16) / 255.0, int(rgb_hex[4:6], 16) / 255.0)
    except Exception: 
        pass
    return None

def process_excel_file(filename: str, contents: bytes):
    wb = None
    try:
        target_tab_name = "新竹車源" if "新竹" in filename else "E車源"
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws_main = wb[wb.sheetnames[0]]
        
        client = get_gspread_client()
        doc = client.open_by_key(SHEET_ID)
        
        try: target_gsheet_main = doc.worksheet(target_tab_name)
        except Exception: return {"status": "error", "message": f"找不到分頁「{target_tab_name}」"}

        data_to_upload_main, color_requests_main = [], []

        if target_tab_name == "新竹車源":
            color_requests_main = [{
                "repeatCell": {
                    "range": { "sheetId": target_gsheet_main.id }, 
                    "cell": {"userEnteredFormat": {"backgroundColor": { "red": 1.0, "green": 1.0, "blue": 1.0 }}}, 
                    "fields": "userEnteredFormat.backgroundColor"
                }
            }]
            
            header_row_idx, headers_main = -1, []
            for i, row in enumerate(ws_main.iter_rows(values_only=True)):
                vals = [str(v).strip() if v is not None else "" for v in row]
                if "車牌" in vals and ("車型" in vals or "品牌" in vals or "廠牌" in vals):
                    header_row_idx = i
                    headers_main = vals
                    break
                    
            if header_row_idx == -1: 
                header_row_idx = 0
                headers_main = [str(cell.value).strip() if cell.value is not None else "" for cell in ws_main[1]]

            old_keys = set()
            is_excel_initial = False
            try:
                old_values = target_gsheet_main.get_all_values()
                if old_values and len(old_values) > header_row_idx:
                    old_hdrs = [str(x).strip() for x in old_values[header_row_idx]]
                    p_idx = old_hdrs.index("車牌") if "車牌" in old_hdrs else -1
                    v_idx = old_hdrs.index("車身") if "車身" in old_hdrs else -1
                    n_idx = old_hdrs.index("新編號") if "新編號" in old_hdrs else -1
                    for row in old_values[header_row_idx+1:]:
                        key = ""
                        if n_idx != -1 and len(row) > n_idx and str(row[n_idx]).strip(): 
                            key = str(row[n_idx]).strip()
                        elif p_idx != -1 and len(row) > p_idx and str(row[p_idx]).strip(): 
                            key = str(row[p_idx]).strip()
                        elif v_idx != -1 and len(row) > v_idx and str(row[v_idx]).strip(): 
                            key = str(row[v_idx]).strip()
                        if key and "車款" not in key and "欄" not in key: 
                            old_keys.add(str(key).replace('.0', '').strip())
                else: 
                    is_excel_initial = True
            except Exception: 
                is_excel_initial = True

            col_model = headers_main.index("車型") if "車型" in headers_main else -1
            plate_idx = headers_main.index("車牌") if "車牌" in headers_main else -1
            vin_idx = headers_main.index("車身") if "車身" in headers_main else -1
            no_idx = headers_main.index("新編號") if "新編號" in headers_main else -1
            year_idx = headers_main.index("年份") if "年份" in headers_main else -1
            col_color = headers_main.index("顏色") if "顏色" in headers_main else -1
            
            if "收訂狀態" not in headers_main: headers_main.append("收訂狀態")
            status_idx = headers_main.index("收訂狀態")
            new_count, new_cars_list = 0, []

            for r_idx, row in enumerate(ws_main.iter_rows()):
                row_values = [cell.value if cell.value is not None else "" for cell in row]
                while len(row_values) < len(headers_main): 
                    row_values.append("")
                    
                if r_idx < header_row_idx or r_idx == header_row_idx or not any(str(v).strip() for v in row_values):
                    data_to_upload_main.append(row_values)
                    target_row_idx = len(data_to_upload_main) - 1
                    for c_idx, cell in enumerate(row):
                        rgb = get_color_rgb(cell)
                        if rgb: 
                            color_requests_main.append({"repeatCell": {"range": { "sheetId": target_gsheet_main.id, "startRowIndex": target_row_idx, "endRowIndex": target_row_idx + 1, "startColumnIndex": c_idx, "endColumnIndex": c_idx + 1 }, "cell": {"userEnteredFormat": {"backgroundColor": { "red": rgb[0], "green": rgb[1], "blue": rgb[2] }}}, "fields": "userEnteredFormat.backgroundColor"}})
                    continue

                while len(row_values) <= status_idx: 
                    row_values.append("")
                    
                n_val = str(row_values[no_idx]).replace('.0', '').strip() if no_idx != -1 else ""
                p_val = str(row_values[plate_idx]).replace('.0', '').strip() if plate_idx != -1 else ""
                v_val = str(row_values[vin_idx]).replace('.0', '').strip() if vin_idx != -1 else ""
                row_key = n_val if n_val else (p_val if p_val else v_val)
                is_subheader = "車款" in p_val or "車輛數" in str(row_values[0]) or "欄1" in str(row_values)
                
                if row_key and not is_subheader and not is_excel_initial and row_key not in old_keys:
                    new_count += 1
                    y_val = str(row_values[year_idx]).replace('.0', '').strip() if year_idx != -1 else ""
                    if len(y_val) == 6 and y_val.isdigit(): y_val = f"{y_val[:4]}年{y_val[4:]}月"
                    elif len(y_val) == 4 and y_val.isdigit(): y_val = f"{y_val}年"
                    m_val = str(row_values[col_model]).strip() if col_model != -1 else ""
                    c_val = str(row_values[col_color]).strip() if col_color != -1 else ""
                    new_cars_list.append(f"{y_val} {m_val} {c_val} #{p_val if p_val else '(無車牌)'}")
                    old_keys.add(row_key)
                
                target_row_idx = len(data_to_upload_main)
                for c_idx, cell in enumerate(row):
                    rgb = get_color_rgb(cell)
                    if rgb: 
                        color_requests_main.append({"repeatCell": {"range": { "sheetId": target_gsheet_main.id, "startRowIndex": target_row_idx, "endRowIndex": target_row_idx + 1, "startColumnIndex": c_idx, "endColumnIndex": c_idx + 1 }, "cell": {"userEnteredFormat": {"backgroundColor": { "red": rgb[0], "green": rgb[1], "blue": rgb[2] }}}, "fields": "userEnteredFormat.backgroundColor"}})
                            
                if not is_subheader:
                    col_i_val = str(row_values[8]).strip() if len(row_values) > 8 else ""
                    row_status = "在庫"
                    if "未售" in col_i_val: row_status = "在庫"
                    elif "收訂" in col_i_val or "客訂" in col_i_val: row_status = "已收訂"
                    elif "售" in col_i_val and "未售" not in col_i_val: row_status = "已售"
                    row_values[status_idx] = row_status

                data_to_upload_main.append(row_values)
                
            try:
                target_gsheet_main.clear()
                target_gsheet_main.update(values=[[str(cell) for cell in row] for row in data_to_upload_main], range_name='A1')
                if header_row_idx > 0 and len(data_to_upload_main) > 1:
                    target_gsheet_main.update_acell('A2', f'="共"&SUMPRODUCT(--(LEN(TRIM($C${header_row_idx + 2}:$C$500))>0))&"台"')
                else: 
                    target_gsheet_main.update_acell('A2', '="共"&SUMPRODUCT(--(LEN(TRIM($C$5:$C$133))>0))&"台"')
                if color_requests_main: 
                    doc.batch_update({"requests": color_requests_main})
            except Exception as e: 
                return {"status": "error", "message": f"新竹寫入失敗：{str(e)}"}

        else:
            headers_main = [str(cell.value).strip() if cell.value is not None else "" for cell in ws_main[1]]
            old_keys = set()
            is_excel_initial = False
            try:
                old_values = target_gsheet_main.get_all_values()
                if old_values and len(old_values) > 1:
                    old_hdrs = [str(x).strip() for x in old_values[0]]
                    p_idx = old_hdrs.index("車牌") if "車牌" in old_hdrs else -1
                    v_idx = old_hdrs.index("車身") if "車身" in old_hdrs else -1
                    n_idx = old_hdrs.index("新編號") if "新編號" in old_hdrs else -1
                    for row in old_values[1:]:
                        key = ""
                        if n_idx != -1 and len(row) > n_idx and str(row[n_idx]).strip(): 
                            key = str(row[n_idx]).strip()
                        elif p_idx != -1 and len(row) > p_idx and str(row[p_idx]).strip(): 
                            key = str(row[p_idx]).strip()
                        elif v_idx != -1 and len(row) > v_idx and str(row[v_idx]).strip(): 
                            key = str(row[v_idx]).strip()
                        if key: old_keys.add(str(key).replace('.0', '').strip())
                else: 
                    is_excel_initial = True
            except Exception: 
                is_excel_initial = True

            col_model = headers_main.index("車型") if "車型" in headers_main else -1
            plate_idx = headers_main.index("車牌") if "車牌" in headers_main else -1
            vin_idx = headers_main.index("車身") if "車身" in headers_main else -1
            no_idx = headers_main.index("新編號") if "新編號" in headers_main else -1
            year_idx = headers_main.index("年份") if "年份" in headers_main else -1
            
            if "狀態" not in headers_main: headers_main.append("狀態")
            status_col_idx = headers_main.index("狀態")
            data_to_upload_main = [headers_main]
            new_count, new_cars_list = 0, []

            for row in ws_main.iter_rows(min_row=2):
                row_values = [cell.value if cell.value is not None else "" for cell in row]
                if not any(str(v).strip() for v in row_values): continue
                while len(row_values) <= status_col_idx: row_values.append("")
                while len(row_values) < len(headers_main): row_values.append("")
                
                n_val = str(row_values[no_idx]).replace('.0', '').strip() if no_idx != -1 else ""
                p_val = str(row_values[plate_idx]).replace('.0', '').strip() if plate_idx != -1 else ""
                v_val = str(row_values[vin_idx]).replace('.0', '').strip() if vin_idx != -1 else ""
                row_key = n_val if n_val else (p_val if p_val else v_val)
                
                if row_key and not is_excel_initial and row_key not in old_keys:
                    new_count += 1
                    y_val = str(row_values[year_idx]).replace('.0', '').strip() if year_idx != -1 else ""
                    if len(y_val) == 6 and y_val.isdigit(): y_val = f"{y_val[:4]}年{y_val[4:]}月"
                    elif len(y_val) == 4 and y_val.isdigit(): y_val = f"{y_val}年"
                    m_val = str(row_values[col_model]).strip() if col_model != -1 else ""
                    new_cars_list.append(f"{y_val} {m_val} #{p_val if p_val else '(無車牌)'}")
                    old_keys.add(row_key)
                
                has_color = False
                for cell in row:
                    if get_color_rgb(cell): has_color = True
                    
                status_val = str(row_values[status_col_idx]).strip()
                if "取證" in status_val: row_values[status_col_idx] = "取證"
                elif "Anti已收訂" in status_val or "已收訂" in status_val: row_values[status_col_idx] = "Anti已收訂" if "Anti" in status_val else "已收訂"
                elif has_color or "已售" in status_val: row_values[status_col_idx] = "已售"
                else:
                    if not status_val: row_values[status_col_idx] = "在庫"
                data_to_upload_main.append(row_values)
                
            target_gsheet_main.clear()
            target_gsheet_main.update(values=[[str(cell) for cell in row] for row in data_to_upload_main], range_name='A1')

        load_and_clean_data()
        msg = "處理成功"
        if new_cars_list: 
            msg += f"\n✨ 新增 {len(new_cars_list)} 台車輛：\n" + "\n".join(new_cars_list[:10])
        return {"status": "success", "message": msg}
        
    except Exception as e: 
        return {"status": "error", "message": f"處理失敗：{str(e)}"}
    finally:
        if wb: wb.close()
        gc.collect()

def get_valid_credentials(force_u=None, force_p=None):
    global cached_valid_u, cached_valid_p
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    login_url = "https://www.jwincar.com.tw/manage/login/index.php"
    data_url = "https://www.jwincar.com.tw/manage/accounting/accounting_car_list.php?stock=all"
    
    if force_u and force_p:
        try:
            session = requests.Session(); session.headers.update(headers)
            session.post(login_url, data={"strID": force_u, "strPW": force_p, "Submit": "送出"})
            if BeautifulSoup(session.get(data_url + "&page=1", timeout=10).text, "html.parser").find("table", {"id": "carTable"}):
                return force_u, force_p
        except Exception: 
            pass
        return None, None

    if cached_valid_u and cached_valid_p:
        try:
            session = requests.Session(); session.headers.update(headers)
            session.post(login_url, data={"strID": cached_valid_u, "strPW": cached_valid_p, "Submit": "送出"})
            if BeautifulSoup(session.get(data_url + "&page=1", timeout=10).text, "html.parser").find("table", {"id": "carTable"}):
                return cached_valid_u, cached_valid_p
        except Exception: 
            pass

    credentials_to_try = []
    try:
        sheet_u, sheet_p = get_or_create_creds()
        credentials_to_try.append((sheet_u, sheet_p))
        for bu, bp in get_backup_credentials_from_sheet():
            if (bu, bp) not in credentials_to_try: 
                credentials_to_try.append((bu, bp))
    except Exception: 
        pass
    
    for test_u, test_p in credentials_to_try:
        try:
            session = requests.Session(); session.headers.update(headers)
            session.post(login_url, data={"strID": test_u, "strPW": test_p, "Submit": "送出"})
            if BeautifulSoup(session.get(data_url + "&page=1", timeout=10).text, "html.parser").find("table", {"id": "carTable"}):
                cached_valid_u, cached_valid_p = test_u, test_p
                try:
                    if test_u != sheet_u or test_p != sheet_p: 
                        update_creds(test_u, test_p)
                except Exception: 
                    pass
                return test_u, test_p
        except Exception: 
            continue
            
    return None, None

# =========================================================================
# 💡 終極核心同步引擎 (雙軌並行抓取，結合前台車牌探測 Spider)
# =========================================================================
def core_sync_car_source(user_id: str, login_user: str, login_pwd: str):
    try:
        session = requests.Session()
        session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
        login_url = "https://www.jwincar.com.tw/manage/login/index.php"
        data_url = "https://www.jwincar.com.tw/manage/accounting/accounting_car_list.php?stock=all"
        session.post(login_url, data={"strID": login_user, "strPW": login_pwd, "Submit": "送出"})
        
        # ----------------------------------------------------
        # 軌道 1：從「收購合約」抓取查定表的 PKey
        # ----------------------------------------------------
        pkey_map = {}
        try:
            contract_url = "https://www.jwincar.com.tw/manage/Contract/p14_contract_purchase_list.php"
            cp, last_c_row = 1, ""
            while cp <= 3000:
                c_res = session.get(f"{contract_url}?page={cp}", timeout=10)
                c_res.encoding = 'utf-8'
                c_soup = BeautifulSoup(c_res.text, "html.parser")
                c_table = c_soup.find("table")
                if not c_table: break
                
                c_rows = c_table.find_all("tr")
                if len(c_rows) <= 1: break
                curr_c_row = c_rows[1].text.strip()
                if curr_c_row == last_c_row: break
                last_c_row = curr_c_row
                
                c_headers = [th.text.strip() for th in c_rows[0].find_all(["th", "td"])]
                p_col = next((i for i, h in enumerate(c_headers) if any(kw in h for kw in ["車牌", "車號", "牌照"])), -1)
                v_col = next((i for i, h in enumerate(c_headers) if any(kw in h for kw in ["車身", "車架", "VIN"])), -1)
                
                for row in c_rows[1:]:
                    tds = row.find_all("td")
                    if not tds: continue
                    row_pkey = ""
                    for td in tds:
                        btn = td.find("input", value=re.compile(r"鑑定|查定|表")) or td.find("input", onclick=re.compile(r"PKey"))
                        if btn and btn.has_attr("onclick"):
                            m = re.search(r'PKey=(\d+)', btn["onclick"])
                            if m: row_pkey = m.group(1); break
                            
                    if row_pkey:
                        if p_col != -1 and p_col < len(tds):
                            txt = tds[p_col].text.strip().upper().replace("-", "")
                            if txt and txt not in ["—", "-", "NAN"]: pkey_map[txt] = row_pkey
                        if v_col != -1 and v_col < len(tds):
                            txt = tds[v_col].text.strip().upper()
                            if txt and txt not in ["—", "-", "NAN"]: pkey_map[txt] = row_pkey
                cp += 1
        except Exception as e: 
            print(f"Contract PKey fetch error: {e}")

        # ----------------------------------------------------
        # 軌道 2：從「在庫車輛清單」抓取基本車輛資料
        # ----------------------------------------------------
        all_cars_dicts = []
        website_headers = []
        page_num, last_first_row = 1, ""
        
        while page_num <= 3000:
            res = session.get(data_url + f"&page={page_num}")
            res.encoding = 'utf-8'
            soup = BeautifulSoup(res.text, "html.parser")
            table = soup.find("table", {"id": "carTable"})
            if not table: break
            
            rows = table.find_all("tr")
            if len(rows) <= 1: break 
            
            current_first_row = rows[1].text.strip()
            if current_first_row == last_first_row: break
            last_first_row = current_first_row

            if page_num == 1: 
                website_headers = [th.text.replace("⇅", "").strip() for th in rows[0].find_all("th")]
                
            for row in rows[1:]:
                tds = row.find_all("td")
                if not tds: continue
                row_dict = {}
                for idx, td in enumerate(tds):
                    if idx < len(website_headers):
                        h = website_headers[idx]
                        if not h or h == "操作": continue
                        val = td.text.strip()
                        if val in ["—", "-"]: val = ""
                        if td.has_attr("title"): val = td["title"].strip()
                        if h == "狀態":
                            if td.find("span", class_=re.compile(r"sold|已售")): val = "已售"
                            elif td.find("span", class_=re.compile(r"stock|在庫")): val = "在庫"
                            elif td.find("span", class_=re.compile(r"deposit|收訂")): val = "已收訂"
                        row_dict[h] = val
                
                row_plate = str(row_dict.get("車牌", "")).strip().upper().replace("-", "")
                row_vin = str(row_dict.get("車身", "")).strip().upper()
                
                all_cars_dicts.append({
                    "row_dict": row_dict,
                    "row_plate": row_plate,
                    "row_vin": row_vin
                })
            page_num += 1

        if len(all_cars_dicts) < 100: 
            return {"status": "error", "message": f"🚨 數據異常熔斷！為保護原始資料庫已自動拒絕寫入。"}

        # ----------------------------------------------------
        # 軌道 3：🚀 官網前台多執行緒掃描器 (自動找分頁 + 黃金線索)
        # ----------------------------------------------------
        frontend_map = {}
        known_plates = set(car["row_plate"] for car in all_cars_dicts if car["row_plate"])
        
        try:
            front_session = requests.Session()
            front_session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
            
            detail_links = set()
            visited_list_pages = set()
            pages_to_visit = [
                "https://www.jwincar.com.tw/p1_buy.php",
                "https://www.jwincar.com.tw/buy_car.php"
            ]
            
            # 第一階段：掃描 Sitemap (最快)
            try:
                sm_res = front_session.get("https://www.jwincar.com.tw/sitemap.xml", timeout=3)
                links = re.findall(r'(p1_buy_detail\.php\?detail_PKey=\d+)', sm_res.text)
                for link in links: detail_links.add("https://www.jwincar.com.tw/" + link)
            except Exception: pass
            
            # 第二階段：蜘蛛爬蟲動態掃描所有分頁
            while pages_to_visit and len(visited_list_pages) < 50:
                current_url = pages_to_visit.pop(0)
                if current_url in visited_list_pages: continue
                visited_list_pages.add(current_url)
                
                try:
                    f_res = front_session.get(current_url, timeout=5)
                    text = f_res.text
                    
                    # 抓取廣告連結
                    links = re.findall(r'(p1_buy_detail\.php\?detail_PKey=\d+)', text)
                    for link in links: detail_links.add("https://www.jwincar.com.tw/" + link)
                    
                    # 自動尋找「下一頁」之類的分頁按鈕
                    page_links = re.findall(r'href=["\']([^"\']*?(?:page|p|nowPage)=\d+[^"\']*)["\']', text, re.IGNORECASE)
                    for pl in page_links:
                        pl_clean = pl.replace("&amp;", "&").lstrip('/')
                        if not pl_clean.startswith('http'):
                            if pl_clean.startswith('?'):
                                full_pl = current_url.split('?')[0] + pl_clean
                            else:
                                full_pl = "https://www.jwincar.com.tw/" + pl_clean
                        else:
                            full_pl = pl_clean
                            
                        if full_pl not in visited_list_pages:
                            pages_to_visit.append(full_pl)
                except Exception: pass
                    
            # 第三階段：進入每台車廣告，執行「黃金線索」匹配
            def fetch_detail(url):
                try:
                    res = front_session.get(url, timeout=5)
                    html = res.text
                    plate_found = ""
                    
                    # 💡 黃金線索：前台網頁寫「提供車身號碼驗證：」，但其實後面接的是「車牌」！
                    m = re.search(r'提供車身號碼驗證[：:]?\s*([A-Za-z0-9]{2,4}[-\s]*[A-Za-z0-9]{2,4})', html)
                    if m:
                        plate_found = m.group(1).replace('-', '').replace(' ', '').strip().upper()
                        
                    if plate_found and plate_found in known_plates:
                        frontend_map[plate_found] = url
                    else:
                        # 備用保險方案：用已經存在的車牌庫去暴力比對廣告內文
                        html_upper = html.upper()
                        for plate in known_plates:
                            if len(plate) >= 4:
                                parts = re.findall(r'[A-Z]+|\d+', plate)
                                dashed_plate = f"{parts[0]}-{parts[1]}" if len(parts) == 2 else plate
                                if dashed_plate in html_upper or plate in html_upper:
                                    frontend_map[plate] = url
                                    break
                except Exception: pass

            if detail_links:
                with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
                    executor.map(fetch_detail, list(detail_links))
                    
        except Exception as e:
            print(f"Frontend scraping error: {e}")

        # ----------------------------------------------------
        # 軌道 4：將所有資料雙向合併並寫入 Google Sheets (只留「連結」欄位)
        # ----------------------------------------------------
        final_cars_list = []
        for car_wrapper in all_cars_dicts:
            row_dict = car_wrapper["row_dict"]
            row_plate = car_wrapper["row_plate"]
            row_vin = car_wrapper["row_vin"]
            
            # 寫入查定表 PKey
            pkey_val = pkey_map.get(row_plate) or pkey_map.get(row_vin) or ""
            row_dict["查定表PKey"] = pkey_val
            
            # 寫入官網廣告網址 (統一只寫到「連結」欄位，保持資料庫乾淨)
            link_url = frontend_map.get(row_plate, "")
            row_dict["連結"] = link_url
            
            final_cars_list.append(row_dict)

        client = get_gspread_client()
        doc = client.open_by_key(SHEET_ID)
        target_gsheet_main = doc.worksheet("E車源")
        existing_data = target_gsheet_main.get_all_values()
        existing_headers = [str(x).strip() for x in existing_data[0]] if existing_data else []

        old_ids = set()
        if len(existing_data) > 1:
            n_idx = existing_headers.index("新編號") if "新編號" in existing_headers else -1
            p_idx = existing_headers.index("車牌") if "車牌" in existing_headers else -1
            for r in existing_data[1:]:
                k = ""
                if n_idx != -1 and len(r) > n_idx and str(r[n_idx]).strip(): k = str(r[n_idx]).strip()
                elif p_idx != -1 and len(r) > p_idx and str(r[p_idx]).strip(): k = str(r[p_idx]).strip()
                if k: old_ids.add(str(k).replace('.0', '').strip())

        df_crawled = pd.DataFrame(final_cars_list)
        
        status_msg = ""
        if "狀態" in df_crawled.columns:
            st_counts = {}
            for st in df_crawled["狀態"]:
                val = str(st).strip() or "在庫"
                st_counts[val] = st_counts.get(val, 0) + 1
            status_parts = [f"{k}: {v}台" for k, v in st_counts.items()]
            if status_parts: status_msg = f"\n📊 狀態分佈：{'、'.join(status_parts)}"

        is_initial = len(old_ids) == 0
        new_count, new_cars_list = 0, []
        if "新編號" in df_crawled.columns or "車牌" in df_crawled.columns:
            for idx, row in df_crawled.iterrows():
                cid = str(row.get("新編號", "")).replace('.0', '').strip() or str(row.get("車牌", "")).replace('.0', '').strip()
                if cid and not is_initial and cid not in old_ids:
                    new_count += 1
                    y = str(row.get("年份", "")).replace('.0', '').strip()
                    if len(y) == 6 and y.isdigit(): y = f"{y[:4]}年{y[4:]}月"
                    elif len(y) == 4 and y.isdigit(): y = f"{y}年"
                    plate = str(row.get('車牌','')).strip() or "(無車牌)"
                    new_cars_list.append(f"{y} {str(row.get('車型','')).strip()} #{plate}")
                    old_ids.add(cid)

        final_headers = list(existing_headers)
        for col in df_crawled.columns:
            if col not in final_headers: final_headers.append(col)
        if not final_headers: final_headers = list(df_crawled.columns)
        
        # 確保功能性欄位都有被建立
        if "查定表PKey" not in final_headers: final_headers.append("查定表PKey")
        if "連結" not in final_headers: final_headers.append("連結")

        df_aligned = df_crawled.reindex(columns=final_headers).fillna("")
        data_to_upload = [final_headers] + df_aligned.values.tolist()
        
        target_gsheet_main.clear()
        target_gsheet_main.update(values=data_to_upload, range_name='A1')
        
        status_col_idx = final_headers.index("狀態") if "狀態" in final_headers else -1
        if status_col_idx != -1:
            try: sold_gsheet = doc.worksheet("E車源售出")
            except Exception: sold_gsheet = doc.add_worksheet(title="E車源售出", rows="1000", cols="30")
            try: old_records = sold_gsheet.get_all_records()
            except Exception: old_records = []
            
            new_records = []
            for row in data_to_upload[1:]:
                if str(row[status_col_idx]).strip() and str(row[status_col_idx]).strip() != "在庫":
                    padded = list(row) + [""] * (len(final_headers) - len(row))
                    new_records.append(dict(zip(final_headers, padded)))
                    
            if new_records or old_records:
                merged_dict = {}
                for rec in old_records:
                    pk = str(rec.get("車牌", "")).strip() or str(rec.get("新編號", "")).strip() or str(rec.get("車身", "")).strip()
                    if pk: merged_dict[pk] = rec
                for rec in new_records:
                    pk = str(rec.get("車牌", "")).strip() or str(rec.get("新編號", "")).strip() or str(rec.get("車身", "")).strip()
                    if pk: merged_dict[pk] = rec
                    else: merged_dict[str(uuid.uuid4())] = rec
                
                sold_headers = list(final_headers)
                for rec in merged_dict.values():
                    for k in rec.keys():
                        if k not in sold_headers and str(k).strip(): sold_headers.append(k)
                final_sold_data = [sold_headers] + [[str(rec.get(h, "")) for h in sold_headers] for rec in merged_dict.values()]
                sold_gsheet.clear()
                sold_gsheet.update(values=final_sold_data, range_name='A1')

        load_and_clean_data()
        msg = f"🤖 更新成功！共抓取 {len(all_cars_dicts)} 筆車源。{status_msg}"
        if new_count > 0:
            msg += f"\n✨ 自動發現 {new_count} 台新車：\n" + "\n".join(new_cars_list[:10])
            if new_count > 10: msg += f"\n...等共 {new_count} 台"
        return {"status": "success", "message": msg}

    except Exception as e: 
        return {"status": "error", "message": f"爬蟲發生錯誤：{str(e)}"}
    finally: 
        gc.collect()

@app.get("/api/view_inspection", response_class=HTMLResponse)
def view_inspection(PKey: str = ""):
    global view_api_session
    if not PKey: return "<h1>❌ 錯誤：缺少 PKey</h1>"
    login_url = "https://www.jwincar.com.tw/manage/login/index.php"
    target_url = f"https://www.jwincar.com.tw/manage/accounting/accounting_car_inspection_view.php?PKey={PKey}"
    
    try:
        if view_api_session is None:
            view_api_session = requests.Session()
            view_api_session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
            u, p = get_valid_credentials()
            if u and p: view_api_session.post(login_url, data={"strID": u, "strPW": p, "Submit": "送出"})
            
        res = view_api_session.get(target_url, timeout=10)
        res.encoding = 'utf-8'
        
        if "login/index.php" in res.text or "請先登入" in res.text or "請輸入密碼" in res.text or "login" in res.url.lower():
            u, p = get_valid_credentials()
            if not u: return "<h1>❌ 錯誤：自動登入失敗，請確認後台密碼。</h1>"
            view_api_session = requests.Session()
            view_api_session.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
            view_api_session.post(login_url, data={"strID": u, "strPW": p, "Submit": "送出"})
            res = view_api_session.get(target_url, timeout=10)
            res.encoding = 'utf-8'
            
        soup = BeautifulSoup(res.text, "html.parser")
        
        for script in soup.find_all("script"):
            if script.string:
                s_code = script.string
                if "location.href" in s_code or "window.location" in s_code or "location.replace" in s_code:
                    s_code = s_code.replace("window.location.href", "console.log")
                    s_code = s_code.replace("window.location", "console.log")
                    s_code = s_code.replace("location.href", "console.log")
                    s_code = s_code.replace("location.replace", "console.log")
                    script.string = s_code
                    
        for meta in soup.find_all("meta", attrs={"http-equiv": re.compile(r"refresh", re.I)}): meta.decompose()
        
        base_tag = soup.new_tag('base', href="https://www.jwincar.com.tw/manage/accounting/")
        if soup.head: soup.head.insert(0, base_tag)
        else: soup.insert(0, base_tag)
        
        style_tag = soup.new_tag('style')
        style_tag.string = "body { background-color: #f3f4f6; } .print-btn { display: none !important; } input[value*='編輯'], input[value*='修改'], .edit-btn, button[id*='edit'], a[href*='edit'] { display: none !important; }"
        if soup.head: soup.head.append(style_tag)
        
        for btn in soup.find_all(["input", "button", "a"]):
            val = str(btn.get("value", "")).strip()
            txt = btn.text.strip()
            if "編輯" in val or "編輯" in txt or "修改" in val or "修改" in txt:
                btn.decompose()
                
        return str(soup)
    except Exception as e:
        return f"<h1>❌ 抓取失敗：網路異常 ({str(e)})</h1>"

@app.get("/api/sync_car_source")
def api_sync_car_source(user_id: str = "", u: str = "", p: str = ""):
    if not check_permission(user_id, "更新車源"): return {"status": "error", "message": "⛔ 權限不足！請聯繫管理員開通「更新車源」權限。"}
    valid_u, valid_p = get_valid_credentials(u, p)
    if not valid_u: return {"status": "need_login", "message": "⚠️ 系統自動嘗試備用密碼失敗，請手動輸入最新的帳號與密碼。"}
    return core_sync_car_source(user_id, valid_u, valid_p)

@app.post("/api/parse_ad")
async def parse_ad(request: Request):
    data = await request.json()
    raw_text = data.get("text", "").strip()
    found_brand, found_model = "", ""
    lines = [l.strip() for l in raw_text.split('\n') if l.strip()]

    for brand in KNOWN_MAKES:
        if brand.lower() in raw_text.lower():
            found_brand = brand
            break

    target_line = ""
    for line in lines:
        if "】" in line or (found_brand and found_brand.lower() in line.lower()):
            target_line = line
            break
            
    if not target_line and lines: 
        target_line = lines[0]

    if target_line:
        clean_line = re.sub(r'【.*?】', '', target_line)
        clean_line = re.sub(r'\d{4}\s*[年式]*', '', clean_line)
        if found_brand: 
            clean_line = re.compile(re.escape(found_brand), re.IGNORECASE).sub("", clean_line)
        found_model = re.sub(r'^[\s\-式年]*', '', clean_line).strip()

    man_date_str = ""
    man_patterns = [
        (r'(20\d{2})\s*年\s*(\d{1,2})\s*月?\s*出廠', lambda m: f"{m.group(1)}年{int(m.group(2))}月"),
        (r'(20\d{2})[^\d]{1,10}(\d{1,2})[^\d]*出廠', lambda m: f"{m.group(1)}年{int(m.group(2))}月"),
        (r'出廠[^\d]{0,10}(20\d{2})[^\d]+(\d{1,2})', lambda m: f"{m.group(1)}年{int(m.group(2))}月"),
        (r'(20\d{2})[^\d]*出廠', lambda m: f"{m.group(1)}年1月"),
        (r'^(20\d{2})年\s', lambda m: f"{m.group(1)}年1月"), 
        (r'(20\d{2})\s*年\s*/?\s*(\d{1,2})\s*月出廠', lambda m: f"{m.group(1)}年{int(m.group(2))}月"),
    ]
    for pat, formatter in man_patterns:
        match = re.search(pat, raw_text, re.MULTILINE)
        if match: 
            man_date_str = formatter(match)
            break

    lic_date_str = ""
    lic_patterns = [
        (r'(20\d{2})[^\d]+(\d{1,2})[^\d]+(\d{1,2})[^\d]*領牌', lambda m: f"{m.group(1)}年{int(m.group(2)):02d}月{int(m.group(3)):02d}日"),
        (r'(20\d{2})[^\d]+(\d{1,2})[^\d]*領牌', lambda m: f"{m.group(1)}年{int(m.group(2)):02d}月"),
        (r'領牌.*?(20\d{2})[^\d]+(\d{1,2})[^\d]+(\d{1,2})', lambda m: f"{m.group(1)}年{int(m.group(2)):02d}月{int(m.group(3)):02d}日"),
        (r'(20\d{2})\s*年\s*/?\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日領牌', lambda m: f"{m.group(1)}年{int(m.group(2)):02d}月{int(m.group(3)):02d}日"),
    ]
    for pat, formatter in lic_patterns:
        match = re.search(pat, raw_text)
        if match: 
            lic_date_str = formatter(match)
            break

    mileage_str = ""
    m = re.search(r'里程[：:]?\s*([0-9,]+)', raw_text)
    if not m: m = re.search(r'([0-9,]+)\s*公里', raw_text)
    if not m: m = re.search(r'([0-9,]+)\s*km', raw_text, re.IGNORECASE)
    if m: mileage_str = f"{m.group(1)}公里"

    clean_price_text = raw_text.replace(',', '')
    new_p = ""
    store_p = ""
    promo_p = ""

    new_p_match = re.search(r'新車.*?([\d.]+)萬', clean_price_text)
    if new_p_match: new_p = new_p_match.group(1)

    store_p_match = re.search(r'店[內面].*?([\d.]+)萬', clean_price_text)
    if store_p_match: store_p = store_p_match.group(1)

    for rg in [r'網路(?:促銷|價).*?([\d.]+)萬', r'(?:網路)?促銷價.*?([\d.]+)萬', r'優惠價.*?([\d.]+)萬', r'折扣.*?([\d.]+)萬', r'最新優惠.*?([\d.]+)萬']:
        pm = re.search(rg, clean_price_text)
        if pm: 
            promo_p = pm.group(1)
            break

    if not store_p and promo_p:
        try: store_p = f"{float(promo_p)+3:.1f}".replace(".0", "")
        except: pass
    if store_p and not promo_p:
        try: promo_p = f"{float(store_p)-3:.1f}".replace(".0", "")
        except: pass

    if not store_p and not promo_p:
        valid_prices = []
        for match_m in re.finditer(r'([\d.]+)萬', clean_price_text):
            try:
                v = float(match_m.group(1))
                if not (5.0 <= v <= 5000.0): continue
                if new_p and abs(v - float(new_p)) < 0.1: continue
                valid_prices.append(v)
            except: pass
        if valid_prices:
            valid_prices.sort()
            promo_p = str(valid_prices[0])
            if len(valid_prices) > 1: store_p = str(valid_prices[1])
            else: store_p = f"{float(promo_p)+3:.1f}".replace(".0", "")

    loan_term = ""
    loan_monthly = ""
    loan_match = re.search(r'月付.*?(\d+)\$?\s*[:/]\s*(\d+)期', clean_price_text)
    if not loan_match: loan_match = re.search(r'\$(\d+)\s*[:/]\s*(\d+)期', clean_price_text)

    if loan_match:
        v1, v2 = loan_match.group(1), loan_match.group(2)
        if int(v1) > 100: loan_monthly, loan_term = v1, v2
        else: loan_term, loan_monthly = v1, v2
    else:
        term_match = re.search(r'(\d+)期', clean_price_text)
        if term_match: loan_term = term_match.group(1)
        monthly_match = re.search(r'月付.*?(\d+)', clean_price_text)
        if monthly_match: loan_monthly = monthly_match.group(1)

    return {"status": "success", "data": {"brand": found_brand, "model": found_model, "man_date": man_date_str, "lic_date": lic_date_str, "mileage": mileage_str, "new_price": new_p, "store_price": store_p, "promo_price": promo_p, "loan_term": loan_term, "loan_monthly": loan_monthly}}

@app.post("/api/export_board")
async def export_board(request: Request):
    try:
        data = await request.json()
        brand = str(data.get("brand", ""))
        model = str(data.get("model", ""))
        
        price_val = str(data.get("price", ""))
        if price_val and "萬" not in price_val:
            price_val += "萬"
        
        template_path = "template.xlsx"
        
        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb["認證表格"] if "認證表格" in wb.sheetnames else wb.active
            
            updates = {
                2: brand,
                3: model,
                4: str(data.get("man_date", "")),
                5: str(data.get("lic_date", "")),
                6: str(data.get("mileage", "")),
                7: price_val
            }
            
            for r, val in updates.items():
                cell = ws.cell(row=r, column=2)
                cell.value = val
                
                if cell.alignment:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        text_rotation=cell.alignment.text_rotation,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=True, 
                        indent=cell.alignment.indent
                    )
                else:
                    cell.alignment = Alignment(shrink_to_fit=True)
                
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "認證表格"
            
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 75
            
            updates_list = [
                (2, "廠牌", brand),
                (3, "車型", model),
                (4, "出廠日期", str(data.get("man_date", ""))),
                (5, "領牌日期", str(data.get("lic_date", ""))),
                (6, "里程數", str(data.get("mileage", ""))),
                (7, "售價", price_val)
            ]
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_label = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            for row_idx, label, val in updates_list:
                ws.row_dimensions[row_idx].height = 70
                
                c_label = ws.cell(row=row_idx, column=1, value=label)
                c_label.font = Font(name='微軟正黑體', size=36, bold=True)
                c_label.alignment = Alignment(horizontal='center', vertical='center')
                c_label.border = thin_border
                c_label.fill = fill_label
                
                c_val = ws.cell(row=row_idx, column=2, value=val)
                color = "FF0000" if label == "售價" else "000000"
                c_val.font = Font(name='微軟正黑體', size=40, bold=True, color=color)
                c_val.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, wrap_text=True)
                c_val.border = thin_border
                
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        year_match = re.search(r'(\d{4})', str(data.get("man_date", "")))
        year_prefix = year_match.group(1) if year_match else "0000"
        safe_model = re.sub(r'[\\/*?:"<>|]', "", model)
        filename = f"{year_prefix}_{brand}_{safe_model}.xlsx"
        
        return StreamingResponse(
            stream, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
        )
    except Exception as e:
        print(f"Export Error: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@app.get("/api/my_permissions")
def get_my_permissions(user_id: str = "", user_name: str = ""):
    if not user_id: return {"status": "error", "message": "ID 缺失"}
    try:
        cached_perms = api_cache.get("perm_dict_list")
        if cached_perms is not None:
            for r in cached_perms:
                if str(r.get("LINE ID", "")).strip() == str(user_id).strip():
                    exp_str = str(r.get("到期日", "")).strip()
                    is_super = str(r.get("最高管理員", "")).strip().upper() == "V"
                    if exp_str and not is_super:
                        try:
                            tw_now = datetime.utcnow() + timedelta(hours=8)
                            exp_date = datetime.strptime(exp_str.replace("-", "/"), "%Y/%m/%d") + timedelta(days=1)
                            if tw_now >= exp_date:
                                for k in r.keys():
                                    if k not in ["LINE ID", "姓名", "到期日"]: r[k] = ""
                        except: pass
                    return {"status": "success", "permissions": r, "is_new": False}
                    
        client = get_gspread_client()
        doc = client.open_by_key(SHEET_ID)
        ws = doc.worksheet("權限管理")
        raw_data = ws.get_all_values()
        if not raw_data: return {"status": "error", "message": "表單為空"}
        headers = raw_data[0]
        records = [dict(zip(headers, row)) for row in raw_data[1:]]
        api_cache.set("perm_dict_list", records)
        
        user_id_clean = str(user_id).strip()
        found_row_index, found_user_data = -1, None
        
        for i, r in enumerate(records):
            if str(r.get("LINE ID", "")).strip() == user_id_clean:
                found_row_index, found_user_data = i + 2, r; break
                
        if found_user_data:
            if user_name and str(found_user_data.get("姓名", "")) != user_name: 
                ws.update_cell(found_row_index, 1, user_name) 
            exp_str = str(found_user_data.get("到期日", "")).strip()
            is_super = str(found_user_data.get("最高管理員", "")).strip().upper() == "V"
            if exp_str and not is_super:
                try:
                    tw_now = datetime.utcnow() + timedelta(hours=8)
                    exp_date = datetime.strptime(exp_str.replace("-", "/"), "%Y/%m/%d") + timedelta(days=1)
                    if tw_now >= exp_date:
                        for k in found_user_data.keys():
                            if k not in ["LINE ID", "姓名", "到期日"]: found_user_data[k] = ""
                except: pass
            return {"status": "success", "permissions": found_user_data, "is_new": False}
            
        ws.append_row([user_name, user_id_clean], value_input_option='USER_ENTERED')
        api_cache.clear("perm_dict_list")
        return {"status": "success", "permissions": {}, "is_new": True}
    except Exception as e: return {"status": "error", "message": str(e)}

@app.get("/api/check_auth")
def check_auth(user_id: str = "", action: str = ""): return {"authorized": check_permission(user_id, action)}

@app.get("/api/refresh")
def refresh_data(): load_and_clean_data(); return {"message": "資料已更新"}

@app.get("/api/options")
def get_options():
    if cached_df is None: load_and_clean_data()
    brands = sorted([str(x) for x in cached_df['廠牌'].unique() if x])
    locations = sorted([str(x) for x in cached_df['車輛位置'].unique() if x])
    props = sorted([str(x) for x in cached_df['filter_property'].unique() if x and x != "其他"])
    if "其他" in cached_df['filter_property'].unique(): props.append("其他")
    return {"brands": ["全部"] + brands, "locations": ["全部"] + locations, "properties": ["全部"] + props}

@app.get("/api/cars")
def get_cars(brand: str = "全部", location: str = "全部", prop: str = "全部", model: str = "", plate: str = "", person: str = "", min_price: float = 0.0, max_price: float = 99999.0, sort_by: str = "預設", limit: int = 100, hide_no_price: str = "false", hide_sold: str = "false", hide_cert: str = "false", hide_reserved: str = "false"):
    if cached_df is None: load_and_clean_data()
    res = cached_df.copy()
    if brand != "全部": res = res[res['廠牌'] == brand]
    if location != "全部": res = res[res['車輛位置'] == location]
    if prop != "全部": res = res[res['filter_property'] == prop]
    if model: res = res[res['車型'].astype(str).str.replace(r'[\s\-]', '', regex=True).str.lower().str.contains(re.sub(r'[\s\-]', '', model).lower(), na=False)]
    if plate: res = res[res['車牌'].astype(str).str.replace(r'[\s\-]', '', regex=True).str.lower().str.contains(re.sub(r'[\s\-]', '', plate).lower(), na=False)]
    if person:
        mask = pd.Series(False, index=res.index)
        if '採購' in res.columns: mask = mask | res['採購'].astype(str).str.contains(person, case=False)
        res = res[mask]
    res = res[(res['顯示價格'] >= min_price) & (res['顯示價格'] <= max_price)]
    if hide_no_price == "true": res = res[res['顯示價格'] > 0]
    if hide_sold == "true": res = res[res['is_sold'] == False]
    if hide_cert == "true": res = res[res['is_cert'] == False]
    if hide_reserved == "true": res = res[res['is_reserved'] == False]
    if sort_by == "價格低到高": res = res.sort_values('顯示價格')
    elif sort_by == "價格高到低": res = res.sort_values('顯示價格', ascending=False)
    elif sort_by == "最新入庫": res = res.sort_values('入庫_dt', ascending=False)
    else: res = res.sort_values('年份', ascending=False)
    return {"total": len(res), "data": res.head(limit).to_dict(orient="records")}

@app.get("/api/search_plate")
def search_plate(plate: str):
    if cached_df is None: load_and_clean_data()
    res = cached_df.copy()
    if '車牌' in res.columns:
        matches = res[res['車牌'].astype(str).str.replace(r'[\s\-]', '', regex=True).str.upper().str.contains(re.sub(r'[\s\-]', '', plate).upper(), na=False)]
        if len(matches) > 0:
            car_data = matches.iloc[0].to_dict()
            match = re.search(r'\d{4}', str(car_data.get('年份', '')))
            car_data['clean_year'] = match.group(0) if match else str(car_data.get('年份', '')).replace('.0', '')
            return {"status": "success", "data": car_data}
    return {"status": "error", "message": "查無此車"}

@app.get("/api/search_car_multi")
def search_car_multi(mode: str = "plate", query: str = ""):
    if cached_df is None: load_and_clean_data()
    res, query = cached_df.copy(), str(query).strip().upper()
    if not query: return {"status": "error", "message": "請輸入關鍵字"}
    matches = pd.DataFrame()
    if mode == "plate" and '車牌' in res.columns: matches = res[res['車牌'].astype(str).str.replace(r'[\s\-]', '', regex=True).str.upper().str.contains(re.sub(r'[\s\-]', '', query), na=False)]
    elif mode == "id":
        id_cols = [c for c in res.columns if "編號" in c or "序號" in c]
        if id_cols:
            mask = pd.Series(False, index=res.index)
            for c in id_cols: mask = mask | (res[c].astype(str).apply(lambda x: x[:-2] if x.endswith('.0') else x) == query)
            matches = res[mask]
    if len(matches) > 0:
        results = []
        for _, row in matches.iterrows():
            car_data = row.to_dict()
            match = re.search(r'\d{4}', str(car_data.get('年份', '')))
            car_data['clean_year'] = match.group(0) if match else str(car_data.get('年份', '')).replace('.0', '')
            n_id = str(car_data.get('新編號', '')).replace('.0', '').strip()
            car_data['stock_id'] = n_id if n_id else str(car_data.get('舊編號', '')).replace('.0', '').strip()
            results.append(car_data)
        return {"status": "success", "data": results}
    return {"status": "error", "message": "查無此車"}

@app.get("/api/simple_data")
def get_simple_data():
    try:
        df_simple = pd.read_csv(SIMPLE_CSV_URL, header=3).dropna(how='all')
        empty_count, new_columns = 0, []
        for c in df_simple.columns:
            if "Unnamed" in str(c) or str(c).strip() == "": empty_count += 1; new_columns.append(f"__未命名_{empty_count}__")
            else: new_columns.append(str(c).strip())
        df_simple.columns = new_columns
        df_simple = df_simple.dropna(axis=1, how='all').fillna("")
        gc.collect() 
        return {"status": "success", "data": df_simple.to_dict(orient="records")}
    except Exception as e: return {"status": "error", "message": f"讀取失敗：{str(e)}"}

@app.get("/api/customers")
def get_customers():
    try:
        client = get_gspread_client()
        sheet = client.open_by_key(SHEET_ID).worksheet("客資紀錄")
        raw_values = sheet.get_all_values()
        if not raw_values or len(raw_values) < 2: return {"status": "success", "data": []}
        headers = [str(h).strip() for h in raw_values[0]]
        records = []
        for row in raw_values[1:]:
            row_data = [str(cell).strip().lstrip("'") for cell in row]
            while len(row_data) < len(headers): row_data.append("")
            records.append(dict(zip(headers, row_data)))
        return {"status": "success", "data": list(reversed(records))}
    except Exception as e: return {"status": "error", "message": str(e)}

@app.post("/api/customers")
async def add_customer(request: Request):
    try:
        data = await request.json()
        tw_time = datetime.utcnow() + timedelta(hours=8)
        date_str = tw_time.strftime("%Y/%m/%d %H:%M")
        phone_str = str(data.get("phone", "")).strip()
        if phone_str.startswith("0"): phone_str = f"'{phone_str}"
        row_data = [date_str, data.get("name", ""), phone_str, data.get("needs", ""), data.get("memo", "")]
        client = get_gspread_client()
        sheet = client.open_by_key(SHEET_ID).worksheet("客資紀錄")
        sheet.append_row(row_data, value_input_option='USER_ENTERED')
        return {"status": "success", "message": "客資已新增"}
    except Exception as e: return {"status": "error", "message": str(e)}

@app.post("/api/upload_excel")
async def upload_excel(file: UploadFile = File(...)):
    filename = file.filename
    contents = await file.read()
    if filename.lower().endswith('.pdf'): return process_pdf_file(filename, contents)
    elif "customer" in filename.lower() or "客資" in filename: return process_crm_excel(filename, contents)
    else: return process_excel_file(filename, contents)

@handler.add(MessageEvent, message=TextMessage)
def handle_text_message(event):
    user_id = event.source.user_id
    text = event.message.text.strip()
    
    if text == "我的ID" or text.lower() == "my id":
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=f"👤 您的 LINE ID 為：\n{user_id}"))
        return

    if text in ["更新車源", "抓取車源"]:
        if not check_permission(user_id, "更新車源"):
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text="⛔ 抱歉，您沒有執行「更新車源」的權限。"))
            return
            
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text="🤖 身份確認！正在連線後台抓取車源..."))
        def run_task():
            try:
                valid_u, valid_p = get_valid_credentials()
                if not valid_u:
                    line_bot_api.push_message(user_id, TextSendMessage(text="🚨 後台密碼已更改，自動嘗試備用密碼也全數失敗。\n請至網頁版手動輸入新密碼！"))
                    return
                res = core_sync_car_source(user_id, valid_u, valid_p)
                line_bot_api.push_message(user_id, TextSendMessage(text=res["message"]))
            except Exception as e:
                line_bot_api.push_message(user_id, TextSendMessage(text=f"❌ 發生錯誤：{str(e)}"))
        threading.Thread(target=run_task).start()
        return

    if text.startswith("客資") or text.startswith("記客"):
        try:
            parts = [p.strip() for p in text.split('/')]
            if len(parts) >= 4:
                name, phone, needs = parts[1], parts[2], parts[3]
                memo = parts[4] if len(parts) > 4 else ""
                phone_val = f"'{phone}" if phone.startswith("0") else phone
                tw_time = (datetime.utcnow() + timedelta(hours=8)).strftime("%Y/%m/%d %H:%M")
                client = get_gspread_client()
                sheet = client.open_by_key(SHEET_ID).worksheet("客資紀錄")
                sheet.append_row([tw_time, name, phone_val, needs, "", "新客詢問", memo], value_input_option='USER_ENTERED')
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text=f"✅ 客資建檔成功！\n姓名：{name}"))
            else:
                line_bot_api.reply_message(event.reply_token, TextSendMessage(text="❌ 格式錯誤！請輸入：\n客資 / 姓名 / 電話 / 需求"))
        except Exception as e:
            line_bot_api.reply_message(event.reply_token, TextSendMessage(text=f"❌ 寫入錯誤：{str(e)}"))
        return

    line_bot_api.reply_message(event.reply_token, TextSendMessage(text="🤖 您好！我是自動小幫手。\n\n▶️ 【車源更新】請說：「更新車源」\n▶️ 【我的權限】請說：「我的ID」\n▶️ 【手動記客】客資 / 姓名 / 電話 / 需求"))

@handler.add(MessageEvent, message=FileMessage)
def handle_file_message(event):
    user_id = event.source.user_id
    message_id = event.message.id
    filename = event.message.file_name
    
    if not check_permission(user_id, "上傳檔案"):
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text="❌ 抱歉，您目前沒有「上傳檔案」的權限。"))
        return

    is_excel = filename.lower().endswith('.xlsx')
    is_pdf = filename.lower().endswith('.pdf')
    
    if not (is_excel or is_pdf):
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text="❌ 請上傳 .xlsx 或是 .pdf 格式的檔案！"))
        return
    
    line_bot_api.reply_message(event.reply_token, TextSendMessage(text="⏳ 權限確認！收到檔案，正在幫您解析資料..."))
    
    def process_and_notify():
        try:
            message_content = line_bot_api.get_message_content(message_id)
            contents = b"".join([chunk for chunk in message_content.iter_content()])
            
            if is_pdf: result = process_pdf_file(filename, contents)
            elif "customer" in filename.lower() or "客資" in filename: result = process_crm_excel(filename, contents)
            else: result = process_excel_file(filename, contents)
                
            if result["status"] == "success": line_bot_api.push_message(user_id, TextSendMessage(text="✅ 處理完成！\n" + result["message"]))
            else: line_bot_api.push_message(user_id, TextSendMessage(text="❌ 處理失敗：\n" + result["message"]))
        except Exception as e: line_bot_api.push_message(user_id, TextSendMessage(text=f"❌ 發生系統錯誤：\n{str(e)}"))
        finally: gc.collect()

    threading.Thread(target=process_and_notify).start()

@app.post("/callback")
async def callback(request: Request):
    signature = request.headers.get("X-Line-Signature", "")
    body = await request.body()
    try: 
        handler.handle(body.decode("utf-8"), signature)
    except InvalidSignatureError: 
        raise HTTPException(status_code=400)
    return "OK"

@app.get("/")
def serve_home(): 
    return FileResponse("index.html")

@app.get("/ping")
def ping(): 
    return {"status": "ok"}

@app.get("/{path}")
def serve_pages(path: str):
    if os.path.exists(f"{path}.html"): 
        return FileResponse(f"{path}.html")
    return FileResponse("index.html")
